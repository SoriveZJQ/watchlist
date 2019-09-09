#!/usr/bin/env python
# -*- coding:utf-8 -*-


import sys
import time
import requests
import re
import hashlib
from openpyxl import load_workbook
from openpyxl import Workbook
from pyquery import PyQuery as pq

#MD5加密
def md5Encode(str1):
    m = hashlib.md5()
    m.update(str1.encode(encoding="utf-8"))
    return m.hexdigest()

#Int转换为Str
def IntToStr(num):
    return str(num)

time_start = None
time_stop = None


LOGIN_PWDS = ['201707030103']
login_url = 'http://202.115.133.173:805/Common/Handler/UserLogin.ashx'
score_url = 'http://202.115.133.173:805/SearchInfo/Score/ScoreList.aspx'
person = {}


n = None
accounts = None
passwords = None
terms = None
errors = None


def query(file, xterm):
    global n
    global accounts
    global passwords
    global terms
    global errors
    wb = load_workbook(file)
    sheet_names = wb.sheetnames
    sheet = wb[sheet_names[0]]
    accounts = []
    passwords = []
    terms = []
    for cell in list(sheet.columns)[0]:
        if cell.value is None:
            continue
        value = str(cell.value)
        accounts.append(value)
        #print(value)
        val1 = int(value[0:4])
        val2 = xterm % 2
        if val2==0:
            val1 = str(val1+int((xterm-1)/2))+'02'
        else:
            val1 = str(val1+int(xterm/2))+'01'
        terms.append(val1)
    for cell in list(sheet.columns)[1]:
        if cell.value is None:
            continue
        value = str(cell.value)
        passwords.append(value)
    n = len(accounts)


    errors = 0
    global person

    zh = {'优': '95', '良': '85', '中': '75', '及格': '65', '不及格': '55'}
    for i in range(n):
        try:
            name = ''
            score = {}
            credit = {}
            isRX = {}
            allScore = {}
            allCredit = {}
            allIsRX = {}
            score2 = {}
            credit2 = {}
            isRX2 = {}
            loop = 0
            while len(score) == 0:
                loop += 1
                sign = int(round((time.time() * 1000)))  # 获取时间戳
                o = md5Encode(passwords[i])
                str2 = accounts[i] + IntToStr(sign) + o  # 加密算法
                pwd = md5Encode(str2)  # MD5加密
                params = {
                    'Action': "Login",
                    'userName': accounts[i],
                    'pwd': pwd,
                    'sign': sign,
                }
                session = requests.Session()
                session.post(login_url, params)
                s = session.get(score_url)
                html = s.text
                doc = pq(html)
                nameTxt = doc('.ico_user.ico_blue').text()
                name = re.search('(.*?)\s', nameTxt, re.S)
                courses = doc('.score_right_infor_list.listUl')
                courses = courses.children()
                for item in courses.items():
                    term = item.children('.floatDiv20').text().strip()
                    if term != '学期':
                        title = item.find('div:nth-child(3)').text().strip()
                        cj = item.find('div:nth-child(6)').text().strip()
                        xf = item.find('div:nth-child(5)').text().strip()
                        bh = item.find('div:nth-child(2)').text().strip()
                        if bh[0] == 'R' and bh[1] == 'X':
                            bh_ = 1
                            title = title + '(选修)'
                        else:
                            bh_ = 0

                        if cj in zh.keys():
                            cj = zh[cj]
                        allScore[title] = float(cj)
                        allCredit[title] = float(xf)
                        allIsRX[title] = bh_
                    if term[0:4] == terms[i][0:4]:
                        title = item.find('div:nth-child(3)').text().strip()
                        cj = item.find('div:nth-child(6)').text().strip()
                        xf = item.find('div:nth-child(5)').text().strip()
                        bh = item.find('div:nth-child(2)').text().strip()
                        if bh[0] == 'R' and bh[1] == 'X':
                            bh_ = 1
                            title = title + '(选修)'
                        else:
                            bh_ = 0

                        if cj in zh.keys():
                            cj = zh[cj]
                        score2[title] = float(cj)
                        credit2[title] = float(xf)
                        isRX2[title] = bh_
                    if term == terms[i]:
                        title = item.find('div:nth-child(3)').text().strip()
                        cj = item.find('div:nth-child(6)').text().strip()
                        xf = item.find('div:nth-child(5)').text().strip()
                        bh = item.find('div:nth-child(2)').text().strip()
                        if bh[0] == 'R' and bh[1] == 'X':
                            bh_ = 1
                            title = title + '(选修)'
                        else:
                            bh_ = 0

                        if cj in zh.keys():
                            cj = zh[cj]
                        score[title] = float(cj)
                        credit[title] = float(xf)
                        isRX[title] = bh_

                allXf = 0
                qh = 0
                allXf2 = 0
                qh2 = 0
                allXf3 = 0
                qh3 = 0
                for item in score.keys():
                    if isRX[item] == 0:
                        allXf += credit[item]
                        qh += (score[item] / 10 - 5) * credit[item]

                jd = qh / allXf
                score['AAA本学期绩点(不含选修)'] = float(jd)

                for item in score2.keys():
                    if isRX2[item] == 0:
                        allXf3 += credit2[item]
                        qh3 += (score2[item] / 10 - 5) * credit2[item]

                jd = qh3 / allXf3
                score['AAB本学年绩点(不含选修)'] = float(jd)

                for item in allScore.keys():
                    if allIsRX[item] == 0:
                        allXf2 += allCredit[item]
                        qh2 += (allScore[item] / 10 - 5) * allCredit[item]

                jd = qh2 / allXf2
                score['AAC总绩点(不含选修)'] = float(jd)

                while loop > 20:
                    raise Exception
            person[name[0].strip()] = score

            #self.file_changed_signal.emit('{} 获取成功！'.format(self.accounts[i]))

        except Exception:
            errors += 1
            #self.file_changed_signal.emit('{} 获取<span style="color: red">失败</span>！'.format(self.accounts[i]))



def store(file):
    global person
    max_courses = []
    for course in person.values():
        max_courses.extend(list(course.keys()))
    max_courses = list(set(max_courses))
    max_courses.sort()
    #wb = Workbook()
    #ws = wb.active
    wb = load_workbook(file)
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[0]]
    wb.remove_sheet(ws)
    wb.create_sheet('成绩')
    ws = wb['成绩']
    ws.append(['姓名'] + max_courses)
    for name, score in person.items():
        name = [name]
        for c in max_courses:
            if c not in score.keys():
                score[c] = ''
        score = dict(sorted(score.items(), key=lambda x: x[0]))
        ws.append(name + list(score.values()))
    ws['B1'] = '本学期绩点(不含选修)'
    ws['C1'] = '本学年绩点(不含选修)'
    ws['D1'] = '总绩点(不含选修)'
    wb.save(file)


def start(fileopen, xterm, filestore):
    query(fileopen, xterm)
    store(filestore)
    global errors

    return "执行完毕，失败{}处！---------- {}".format(errors, time.strftime('%H:%M:%S', time.localtime(time.time())))
